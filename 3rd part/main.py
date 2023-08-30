import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
from openpyxl import Workbook, load_workbook


class ZNAK:
    def __init__(self, surname, name, zodiac, birthdate):
        self.surname = surname
        self.name = name
        self.zodiac = zodiac
        self.birthdate = birthdate


class App:
    def __init__(self):
        self.znak_list = []
        self.filtered_znak_list = []
        self.zodiac_signs = ["Телец", "Весы", "Близнецы", "Овен", "Дева", "Рак", "Лев", "Скорпион", "Стрелец", "Козерог"
                             , "Водолей", "Рыбы"]

        self.window = tk.Tk()
        self.window.title("ZNAK Collection")
        self.window.configure(bg='Gray')

        # Создаем виджеты
        self.input_frame = tk.Frame(self.window, bg='Gray')
        self.output_frame = tk.Frame(self.window, bg='Gray')
        self.filter_frame = tk.Frame(self.window, bg='Gray')

        self.input_surname_label = tk.Label(self.input_frame, text="Фамилия:")
        self.input_surname_entry = tk.Entry(self.input_frame)

        self.input_name_label = tk.Label(self.input_frame, text="Имя:")
        self.input_name_entry = tk.Entry(self.input_frame)

        self.input_zodiac_label = tk.Label(self.input_frame, text="Знак Зодиака:")
        self.input_zodiac_entry = tk.Entry(self.input_frame)

        self.input_birthdate_label = tk.Label(self.input_frame, text="Дата рождения (ДД.ММ.ГГГГ):")
        self.input_birthdate_entry = tk.Entry(self.input_frame)

        self.add_button = tk.Button(self.input_frame, text="Добавить", command=self.add_znak)

        self.load_button_txt = tk.Button(self.output_frame, text="Загрузить из txt-файла",
                                         command=self.load_from_file_txt)
        self.save_button_txt = tk.Button(self.output_frame, text="Сохранить в txt-файл", command=self.save_to_file_txt)

        self.load_button_excel = tk.Button(self.output_frame, text="Загрузить из Excel-файла",
                                           command=self.load_from_file_excel)
        self.save_button_excel = tk.Button(self.output_frame, text="Сохранить в Excel-файл",
                                           command=self.save_to_file_excel)

        self.sort_label = tk.Label(self.filter_frame, text="Сортировка:")
        self.sort_var = tk.StringVar()
        self.sort_var.set("Дата рождения")
        self.sort_dropdown = tk.OptionMenu(self.filter_frame, self.sort_var, "Фамилия", "Знак Зодиака", "Дата рождения")
        self.sort_button = tk.Button(self.filter_frame, text="Отсортировать",
                                     command=self.sort_and_update_output_listbox)

        self.search_label = tk.Label(self.filter_frame, text="Поиск по фамилии:")
        self.search_entry = tk.Entry(self.filter_frame)
        self.search_button = tk.Button(self.filter_frame, text="Найти", command=self.filter_znak)

        self.output_listbox = tk.Listbox(self.output_frame)
        self.output_listbox.bind("<<ListboxSelect>>", self.show_selected_znak)

        self.output_info_label = tk.Label(self.output_frame, text="Информация о человеке:")
        self.output_info_text = tk.Text(self.output_frame, height=5)

        self.delete_button = tk.Button(self.output_frame, text="Удалить", command=self.delete_znak)

        # Размещаем виджеты на окне
        self.input_frame.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH)
        self.output_frame.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)
        self.filter_frame.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH)

        self.input_surname_label.pack(side=tk.LEFT)
        self.input_surname_entry.pack(side=tk.LEFT, padx=3)

        self.input_name_label.pack(side=tk.LEFT)
        self.input_name_entry.pack(side=tk.LEFT, padx=3)

        self.input_zodiac_label.pack(side=tk.LEFT)
        self.input_zodiac_entry.pack(side=tk.LEFT, padx=3)

        self.input_birthdate_label.pack(side=tk.LEFT)
        self.input_birthdate_entry.pack(side=tk.LEFT, padx=3)

        self.add_button.pack(side=tk.LEFT, padx=10)

        self.load_button_txt.pack(side=tk.LEFT, padx=10)
        self.save_button_txt.pack(side=tk.LEFT, padx=10)
        self.load_button_excel.pack(side=tk.LEFT, padx=10)
        self.save_button_excel.pack(side=tk.LEFT, padx=10)

        self.sort_label.pack(side=tk.LEFT)
        self.sort_dropdown.pack(side=tk.LEFT, padx=3)
        self.sort_button.pack(side=tk.LEFT, padx=10)

        self.search_label.pack(side=tk.LEFT)
        self.search_entry.pack(side=tk.LEFT, padx=3)
        self.search_button.pack(side=tk.LEFT, padx=10)

        self.output_listbox.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)
        self.output_info_label.pack(side=tk.TOP, padx=10, pady=10)
        self.output_info_text.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.delete_button.pack(side=tk.LEFT, padx=10)

        self.window.mainloop()

    def add_znak(self):
        # Получаем данные из полей ввода
        surname = self.input_surname_entry.get().strip()
        name = self.input_name_entry.get().strip()
        zodiac = self.input_zodiac_entry.get().strip()
        birthdate_str = self.input_birthdate_entry.get().strip()

        # Проверяем, что все поля заполнены
        if not surname or not name or not zodiac or not birthdate_str:
            messagebox.showerror("Ошибка", "Все поля должны быть заполнены!")
            return

        if zodiac not in self.zodiac_signs:
            messagebox.showerror("Ошибка", "Неверный ввод знака зодиака!")
            return

        if not surname.isalpha():
            messagebox.showerror("Ошибка", "Фамилия должна состоять из букв!")
            return

        if not name.isalpha():
            messagebox.showerror("Ошибка", "Имя должна состоять из букв!")
            return

        # Парсим дату рождения
        try:
            day, month, year = map(int, birthdate_str.split("."))
            birthdate = (day, month, year)
            if 1 <= birthdate[2] <= 2023:
                if 1 <= birthdate[1] <= 12:
                    if birthdate[1] == 2:
                        if not 1 <= birthdate[0] <= 28:
                            messagebox.showerror("Ошибка",
                                                 "Неверный формат даты рождения, так как в этом месяце 28 дней")
                            return
                    elif birthdate[1] == 4 or birthdate[1] == 6 or birthdate[1] == 9 or birthdate[1] == 11:
                        if not 1 <= birthdate[0] <= 30:
                            messagebox.showerror("Ошибка",
                                                 "Неверный формат даты рождения, так как в этом месяце 30 дней")
                            return
                    else:
                        if not 1 <= birthdate[0] <= 31:
                            messagebox.showerror("Ошибка",
                                                 "Неверный формат даты рождения, так как в этом месяце 31 день")
                            return
            else:
                messagebox.showerror("Ошибка",
                                     "Неверный формат года")
                return
        except:
            messagebox.showerror("Ошибка", "Неверный формат даты рождения!")
            return

        # Создаем экземпляр класса ZNAK и добавляем его в список
        znak = ZNAK(surname, name, zodiac, birthdate)
        self.znak_list.append(znak)

        # Обновляем список на форме и очищаем поля ввода
        self.sort_and_update_output_listbox()
        self.input_surname_entry.delete(0, tk.END)
        self.input_name_entry.delete(0, tk.END)
        self.input_zodiac_entry.delete(0, tk.END)
        self.input_birthdate_entry.delete(0, tk.END)

    def sort_and_update_output_listbox(self):
        selected_item = self.output_listbox.curselection()
        if not selected_item:
            if len(self.znak_list) == 0:
                messagebox.showerror("Ошибка", "Нет данных для сортировки!")
                return
        # Сортируем список
        sort_field = self.sort_var.get()
        if sort_field == "Фамилия":
            self.znak_list.sort(key=lambda znak: znak.surname)
        elif sort_field == "Знак Зодиака":
            self.znak_list.sort(key=lambda znak: znak.zodiac)
        elif sort_field == "Дата рождения":
            self.znak_list.sort(key=lambda znak: znak.birthdate)

        # Очищаем список
        self.output_listbox.delete(0, len(self.znak_list))

        # Заполняем список
        for i, znak in enumerate(self.znak_list):
            self.output_listbox.insert(tk.END,
                                       f"{i + 1}. {znak.surname} {znak.name} ({znak.zodiac}, {znak.birthdate[0]:02d}."
                                       f"{znak.birthdate[1]:02d}.{znak.birthdate[2]:04d})")

        # Обновляем информационную метку
        self.update_output_info_label()

    def delete_znak(self):
        # Получаем выбранный элемент из списка
        selected_index = self.output_listbox.curselection()
        if not selected_index:
            messagebox.showerror("Ошибка", "Не был выбрал элемент для удаления или элементов вовсе нет!")
            return

        # Удаляем элемент из списка
        index = selected_index[0]
        del self.znak_list[index]

        # Обновляем список на форме
        self.sort_and_update_output_listbox()

    def filter_znak(self):

        check = True
        for i in range(len(self.znak_list)):
            if self.znak_list[i].surname == self.search_entry.get().strip():
                check = False
        if self.search_entry.get().strip() == "":
            if len(self.znak_list) == 0:
                messagebox.showerror("Ошибка", "Вы ничего не ввели!")
                return
            else:
                check = False
        if check:
            messagebox.showerror("Ошибка", "Не существует человека с заданной фамилией!")
            return

        # Получаем строку для фильтрации
        filter_str = self.search_entry.get().strip()
        # Очищаем список
        self.output_listbox.delete(0, tk.END)

        # Заполняем список отфильтрованными элементами
        for i, znak in enumerate(self.znak_list):
            if filter_str.lower() in f"{znak.surname} {znak.name} {znak.zodiac} {znak.birthdate[0]:02d}." \
                                     f"{znak.birthdate[1]:02d}.{znak.birthdate[2]:04d}".lower():
                self.output_listbox.insert(tk.END,
                                           f"{i + 1}. {znak.surname} {znak.name} ({znak.zodiac}, "
                                           f"{znak.birthdate[0]:02d}.{znak.birthdate[1]:02d}.{znak.birthdate[2]:04d})")

        # Обновляем информационную метку
        self.update_output_info_label()

    def update_output_info_label(self):
        # Получаем количество элементов в списке
        count = len(self.znak_list)

        # Получаем количество элементов в отфильтрованном списке
        filter_str = self.search_entry.get().strip()
        filtered_count = sum(1 for znak in self.znak_list if
                             filter_str.lower() in f"{znak.surname} {znak.name} {znak.zodiac} {znak.birthdate[0]:02d}."
                                                   f"{znak.birthdate[1]:02d}.{znak.birthdate[2]:04d}".lower())

        # Обновляем информационную метку
        self.output_info_text.delete("1.0", tk.END)
        self.output_info_text.insert(tk.END, f"Всего записей: {count}\nОтфильтровано записей: {filtered_count}")

    def load_from_file_excel(self):
        filepath = filedialog.askopenfilename(title="Выберите Excel-файл", filetypes=[("Excel files", "*.xlsx")])

        if not filepath:
            return

        try:
            workbook = load_workbook(filepath)
            worksheet = workbook.active
        except:
            messagebox.showerror("Ошибка", "Невозможно открыть файл!")
            return

        for row in worksheet.iter_rows(min_row=1):
            surname = row[0].value.strip()
            name = row[1].value.strip()
            zodiac = row[2].value.strip()
            birthdate_string = row[3].value.strip()

            try:
                day, month, year = map(int, birthdate_string.split("."))
                birthdate = (day, month, year)
            except:
                continue

            znak = ZNAK(surname, name, zodiac, birthdate)
            self.znak_list.append(znak)

        self.sort_and_update_output_listbox()

    def load_from_file_txt(self):
        filename = filedialog.askopenfilename(title="Выберите .txt-файл", filetypes=[("Text files", "*.txt")])
        if filename:
            try:
                with open(filename, "r", encoding="utf-8") as f:
                    # Очищаем список перед загрузкой новых данных
                    self.znak_list.clear()
                    # Читаем данные построчно из файла
                    for line in f:
                        # Извлекаем данные из строки
                        surname, name, zodiac, birthdate_str = line.strip().split(",")
                        day, month, year = map(int, birthdate_str.split("."))
                        birthdate = (day, month, year)
                        if 1 <= birthdate[2] <= 2023:
                            if 1 <= birthdate[1] <= 12:
                                if birthdate[1] == 2:
                                    if not 1 <= birthdate[0] <= 28:
                                        messagebox.showerror("Ошибка",
                                                             "Неверный формат даты рождения, так как в этом месяце "
                                                             "28 дней")
                                        return
                                elif birthdate[1] == 4 or birthdate[1] == 6 or birthdate[1] == 9 or birthdate[1] == 11:
                                    if not 1 <= birthdate[0] <= 30:
                                        messagebox.showerror("Ошибка",
                                                             "Неверный формат даты рождения, так как в этом месяце "
                                                             "30 дней")
                                        return
                                else:
                                    if not 1 <= birthdate[0] <= 31:
                                        messagebox.showerror("Ошибка",
                                                             "Неверный формат даты рождения, так как в этом месяце "
                                                             "31 день")
                                        return
                        else:
                            messagebox.showerror("Ошибка",
                                                 "Неверный формат года")
                            return
                        # Создаем экземпляр класса ZNAK и добавляем его в список
                        znak = ZNAK(surname, name, zodiac, birthdate)
                        self.znak_list.append(znak)
                    # Обновляем список на форме
                    self.sort_and_update_output_listbox()
            except FileNotFoundError:
                messagebox.showerror("Ошибка", f"Файл '{filename}' не найден!")
            except:
                messagebox.showerror("Ошибка", "Не удалось загрузить данные из файла!")

    def save_to_file_excel(self):
        if len(self.znak_list) == 0:
            messagebox.showerror("Ошибка", "Нет данных для сохранения")
            return
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if filepath:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["Фамилия", "Имя", "Знак зодиака", "Дата рождения (ДД.ММ.ГГГГ)"])

            for i, znak in enumerate(self.znak_list):
                worksheet.cell(row=i + 2, column=1, value=znak.surname)
                worksheet.cell(row=i + 2, column=2, value=znak.name)
                worksheet.cell(row=i + 2, column=3, value=znak.zodiac)
                worksheet.cell(row=i + 2, column=4,
                               value=f"{znak.birthdate[0]:02d}.{znak.birthdate[1]:02d}.{znak.birthdate[2]:04d}")

            workbook.save(filepath)

            messagebox.showinfo("Сохранение", "Данные сохранены успешно!")

    def save_to_file_txt(self):
        filename = filedialog.asksaveasfilename(defaultextension=".txt",
                                                filetypes=[("Text Files", "*.txt")])
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as file:
                    for znak in self.znak_list:
                        line = f"{znak.surname},{znak.name},{znak.zodiac},{znak.birthdate[0]}.{znak.birthdate[1]}" \
                               f".{znak.birthdate[2]}\n"
                        file.write(line)
                    messagebox.showinfo("Успех", "Данные успешно сохранены в файл!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при сохранении файла: {e}")

    def show_selected_znak(self, event):
        # Получаем выбранный элемент списка
        selection = self.output_listbox.curselection()
        if not selection:
            return
        index = selection[0]

        # Получаем экземпляр класса ZNAK из списка
        znak = self.znak_list[index]

        # Обновляем информацию на форме
        self.output_info_label.config(text=f"{znak.surname} {znak.name}")
        self.output_info_text.config(state=tk.NORMAL)
        self.output_info_text.delete("1.0", tk.END)
        self.output_info_text.insert(tk.END, f"Знак Зодиака: {znak.zodiac}\n")
        self.output_info_text.insert(tk.END,
                                     f"Дата рождения: {znak.birthdate[0]:02d}.{znak.birthdate[1]:02d}."
                                     f"{znak.birthdate[2]}\n")
        self.output_info_text.config(state=tk.DISABLED)


app = App()
